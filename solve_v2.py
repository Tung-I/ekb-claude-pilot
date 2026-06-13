
from openpyxl import load_workbook
from collections import deque

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

ROWS = ws.max_row
COLS = ws.max_column

cell_color = {}
start = None
end = None
for row in ws.iter_rows():
    for cell in row:
        r = cell.row - 1
        c = cell.column - 1
        fill = cell.fill
        try:
            rgb = str(fill.fgColor.rgb)
            if fill.fill_type == "solid" and len(rgb) == 8:
                color = rgb
            else:
                color = "00000000"
        except Exception:
            color = "00000000"
        cell_color[(r,c)] = color
        if cell.value == "START":
            start = (r, c)
        elif cell.value == "END":
            end = (r, c)

BLUE = "FF0099FF"

def is_blue(r, c):
    return cell_color.get((r,c), "00000000") == BLUE

DIRS = [(-1,0,"UP"), (1,0,"DOWN"), (0,-1,"LEFT"), (0,1,"RIGHT")]
OPPOSITE = {"UP":"DOWN", "DOWN":"UP", "LEFT":"RIGHT", "RIGHT":"LEFT"}

# Try: no backward rule at all (only no revisiting), check intermediate
def bfs_solve_v2(check_intermediate, max_turns=50):
    # State: (r, c, frozenset of visited cells)
    # We DON'T track last_dir
    queue = deque()
    init_visited = frozenset([start])
    queue.append((start[0], start[1], init_visited, [start]))
    seen = set()
    solutions = []
    processed = 0
    while queue:
        r, c, visited, path = queue.popleft()
        processed += 1
        if processed % 500000 == 0:
            print(f"  processed={processed}, queue={len(queue)}, path_len={len(path)}")
        if len(path) > max_turns + 1:
            continue
        state_key = (r, c, visited)
        if state_key in seen:
            continue
        seen.add(state_key)
        for dr, dc, dname in DIRS:
            mid_r, mid_c = r + dr, c + dc
            new_r, new_c = r + 2*dr, c + 2*dc
            if not (0 <= mid_r < ROWS and 0 <= mid_c < COLS):
                continue
            if not (0 <= new_r < ROWS and 0 <= new_c < COLS):
                continue
            if check_intermediate and is_blue(mid_r, mid_c):
                continue
            if is_blue(new_r, new_c):
                continue
            if (new_r, new_c) in visited:
                continue
            new_visited = visited | frozenset([(new_r, new_c)])
            new_path = path + [(new_r, new_c)]
            if (new_r, new_c) == end:
                solutions.append(new_path)
                cells = [f"{chr(65+cc)}{rr+1}" for rr,cc in new_path]
                print(f"SOLUTION ({len(new_path)-1} turns): {cells}")
                if len(solutions) >= 3:
                    print(f"Total processed: {processed}")
                    return solutions
            else:
                queue.append((new_r, new_c, new_visited, new_path))
    print(f"Total processed: {processed}")
    return solutions

print("=== no direction restriction, check intermediate ===")
sols = bfs_solve_v2(True, 50)
if not sols:
    print("\n=== no direction restriction, NO intermediate check ===")
    sols = bfs_solve_v2(False, 50)

if sols:
    path = sols[0]
    print("\nPath details:")
    for j, (r, c) in enumerate(path):
        clr = cell_color[(r,c)]
        print(f"  Turn {j}: {chr(65+c)}{r+1} color={clr}")
    if len(path) >= 12:
        r11, c11 = path[11]
        clr = cell_color[(r11,c11)]
        print(f"\nTURN 11 ANSWER: {chr(65+c11)}{r11+1}, color={clr[2:]}")
    else:
        print(f"Path only {len(path)-1} turns long!")
else:
    print("NO SOLUTION FOUND")
