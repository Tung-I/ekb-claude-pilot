
from openpyxl import load_workbook

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

ROWS = ws.max_row
COLS = ws.max_column

cell_color = {}
start = None
end = None
for row in ws.iter_rows():
    for cell in row:
        r = cell.row - 1
        c = cell.column - 1
        fill = cell.fill
        try:
            rgb = str(fill.fgColor.rgb)
            if fill.fill_type == "solid" and len(rgb) == 8:
                color = rgb
            else:
                color = "00000000"
        except Exception:
            color = "00000000"
        cell_color[(r,c)] = color
        if cell.value == "START":
            start = (r, c)
        elif cell.value == "END":
            end = (r, c)

BLUE = "FF0099FF"

def is_blue(r, c):
    return cell_color.get((r,c), "00000000") == BLUE

DIRS = [(-1,0,"UP"), (1,0,"DOWN"), (0,-1,"LEFT"), (0,1,"RIGHT")]

# Check all non-blue cells and their neighbors
non_blue = [(r,c) for r in range(ROWS) for c in range(COLS) if not is_blue(r,c)]
print(f"Non-blue cells: {len(non_blue)}")

for r, c in non_blue:
    moves = []
    for dr, dc, dname in DIRS:
        mid_r, mid_c = r+dr, c+dc
        new_r, new_c = r+2*dr, c+2*dc
        if not (0 <= mid_r < ROWS and 0 <= mid_c < COLS): continue
        if not (0 <= new_r < ROWS and 0 <= new_c < COLS): continue
        if is_blue(mid_r, mid_c): continue
        if is_blue(new_r, new_c): continue
        moves.append(f"{dname}->{chr(65+new_c)}{new_r+1}")
    if moves:
        print(f"  {chr(65+c)}{r+1}: {moves}")
    else:
        print(f"  {chr(65+c)}{r+1}: DEAD END")
