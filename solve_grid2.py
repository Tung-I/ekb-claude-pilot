
from openpyxl import load_workbook
from collections import deque

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

ROWS = ws.max_row
COLS = ws.max_column

cell_color = {}
start = None
end = None
for row in ws.iter_rows():
    for cell in row:
        r = cell.row - 1
        c = cell.column - 1
        fill = cell.fill
        try:
            rgb = str(fill.fgColor.rgb)
            if fill.fill_type == "solid" and len(rgb) == 8:
                color = rgb
            else:
                color = "00000000"
        except Exception:
            color = "00000000"
        cell_color[(r,c)] = color
        if cell.value == "START":
            start = (r, c)
        elif cell.value == "END":
            end = (r, c)

BLUE = "FF0099FF"

def is_blue(r, c):
    return cell_color.get((r,c), "00000000") == BLUE

DIRS = [(-1,0,"UP"), (1,0,"DOWN"), (0,-1,"LEFT"), (0,1,"RIGHT")]
OPPOSITE = {"UP":"DOWN", "DOWN":"UP", "LEFT":"RIGHT", "RIGHT":"LEFT"}

def bfs_solve(allow_backward, check_intermediate, max_turns=40):
    queue = deque()
    queue.append((start[0], start[1], None, [start]))
    seen = {}
    solutions = []
    processed = 0
    while queue:
        r, c, last_dir, path = queue.popleft()
        processed += 1
        if processed % 100000 == 0:
            print(f"  processed={processed}, queue={len(queue)}, path_len={len(path)}")
        if len(path) > max_turns + 1:
            continue
        key = (r, c, last_dir, frozenset(path))
        if key in seen:
            continue
        seen[key] = True
        path_set = set(path)
        for dr, dc, dname in DIRS:
            if not allow_backward and last_dir and dname == OPPOSITE[last_dir]:
                continue
            mid_r, mid_c = r + dr, c + dc
            new_r, new_c = r + 2*dr, c + 2*dc
            if not (0 <= mid_r < ROWS and 0 <= mid_c < COLS):
                continue
            if not (0 <= new_r < ROWS and 0 <= new_c < COLS):
                continue
            if check_intermediate and is_blue(mid_r, mid_c):
                continue
            if is_blue(new_r, new_c):
                continue
            if (new_r, new_c) in path_set:
                continue
            new_path = path + [(new_r, new_c)]
            if (new_r, new_c) == end:
                solutions.append(new_path)
                cells = [f"{chr(65+cc)}{rr+1}" for rr,cc in new_path]
                print(f"SOLUTION ({len(new_path)-1} turns): {cells}")
                if len(solutions) >= 3:
                    return solutions
            else:
                queue.append((new_r, new_c, dname, new_path))
    print(f"Total processed: {processed}")
    return solutions

print("=== no backward, check intermediate, max 40 turns ===")
sols = bfs_solve(False, True, 40)
if not sols:
    print("\n=== allow backward, check intermediate, max 40 turns ===")
    sols = bfs_solve(True, True, 40)

if sols:
    path = sols[0]
    print("\nPath details:")
    for j, (r, c) in enumerate(path):
        clr = cell_color[(r,c)]
        print(f"  Turn {j}: {chr(65+c)}{r+1} color={clr}")
    if len(path) >= 12:
        r11, c11 = path[11]
        clr = cell_color[(r11,c11)]
        print(f"\nTURN 11 ANSWER: {chr(65+c11)}{r11+1}, color={clr[2:]}")
    else:
        print(f"Path only {len(path)-1} turns long!")
else:
    print("NO SOLUTION FOUND")
