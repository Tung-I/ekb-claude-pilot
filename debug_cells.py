
from openpyxl import load_workbook

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

print("All cell values:")
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            fill = cell.fill
            try:
                rgb = str(fill.fgColor.rgb)
            except:
                rgb = "ERROR"
            print(f"  {cell.coordinate}: value={repr(cell.value)}, fill_type={fill.fill_type}, rgb={rgb}")

print()
print("Sheet names:", wb.sheetnames)
print("Active sheet:", ws.title)
print("Max row:", ws.max_row, "Max col:", ws.max_column)
