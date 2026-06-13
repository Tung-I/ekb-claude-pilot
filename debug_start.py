
from openpyxl import load_workbook

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

ROWS = ws.max_row
COLS = ws.max_column

cell_color = {}
start = None
end = None
for row in ws.iter_rows():
    for cell in row:
        r = cell.row - 1
        c = cell.column - 1
        fill = cell.fill
        try:
            rgb = str(fill.fgColor.rgb)
            if fill.fill_type == "solid" and len(rgb) == 8:
                color = rgb
            else:
                color = "00000000"
        except Exception:
            color = "00000000"
        cell_color[(r,c)] = color
        if cell.value == "START":
            start = (r, c)
        elif cell.value == "END":
            end = (r, c)

BLUE = "FF0099FF"

def is_blue(r, c):
    return cell_color.get((r,c), "00000000") == BLUE

print(f"START={start}, END={end}")
print(f"A1 color: {cell_color.get((0,0))}")
print(f"A2 color: {cell_color.get((1,0))}")
print(f"A3 color: {cell_color.get((2,0))}")
print(f"A1 is_blue: {is_blue(0,0)}")
print(f"A2 is_blue: {is_blue(1,0)}")
print(f"A3 is_blue: {is_blue(2,0)}")

# Check all possible moves from A1 (row=0, col=0)
r, c = 0, 0
DIRS = [(-1,0,"UP"), (1,0,"DOWN"), (0,-1,"LEFT"), (0,1,"RIGHT")]
print("\nPossible moves from A1:")
for dr, dc, dname in DIRS:
    mid_r, mid_c = r+dr, c+dc
    new_r, new_c = r+2*dr, c+2*dc
    in_bounds_mid = 0 <= mid_r < ROWS and 0 <= mid_c < COLS
    in_bounds_new = 0 <= new_r < ROWS and 0 <= new_c < COLS
    mid_blue = is_blue(mid_r, mid_c) if in_bounds_mid else "OOB"
    new_blue = is_blue(new_r, new_c) if in_bounds_new else "OOB"
    mid_cell = f"{chr(65+mid_c)}{mid_r+1}" if in_bounds_mid else "OOB"
    new_cell = f"{chr(65+new_c)}{new_r+1}" if in_bounds_new else "OOB"
    print(f"  {dname}: mid={mid_cell}(blue={mid_blue}), land={new_cell}(blue={new_blue})")
    if in_bounds_mid and in_bounds_new and not mid_blue and not new_blue:
        print(f"    -> VALID MOVE to {new_cell}")
