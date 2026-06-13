import openpyxl
from openpyxl import load_workbook
from collections import deque

wb = load_workbook('/work/pi_rsitaram_umass_edu/tungi/ekb-claude-pilot/data/gaia/raw/65afbc8a-89ca-4ad5-8d62-355bb401f61d.xlsx')
ws = wb.active

ROWS = ws.max_row
COLS = ws.max_column

cell_color = {}
start = None
end = None
for row in ws.iter_rows():
    for cell in row:
        r = cell.row - 1
        c = cell.column - 1
        fill = cell.fill
        if fill and fill.fgColor and fill.fill_type == "solid":
            color = fill.fgColor.rgb
        else:
            color = "00000000"
        cell_color[(r,c)] = color
        if cell.value == "START":
            start = (r, c)
        elif cell.value == "END":
            end = (r, c)

print(f"START: row={start[0]+1}, col={chr(65+start[1])}")
print(f"END: row={end[0]+1}, col={chr(65+end[1])}")

BLUE = "FF0099FF"

def is_blue(r, c):
    return cell_color.get((r,c), "00000000") == BLUE

print("Grid:")
for r in range(ROWS):
    row_str = ""
    for c in range(COLS):
        if (r,c) == start:
            row_str += "S"
        elif (r,c) == end:
            row_str += "E"
        elif is_blue(r, c):
            row_str += "X"
        else:
            clr = cell_color[(r,c)]
            if clr == "FF92D050": row_str += "G"
            elif clr == "FFFFFF00": row_str += "Y"
            elif clr == "FFF478A7": row_str += "P"
            else: row_str += "."
    print(f"  Row {r+1:2d}: {row_str}")

DIRS = [(-1,0,"UP"), (1,0,"DOWN"), (0,-1,"LEFT"), (0,1,"RIGHT")]
OPPOSITE = {"UP":"DOWN", "DOWN":"UP", "LEFT":"RIGHT", "RIGHT":"LEFT"}

def bfs_solve(allow_backward, check_intermediate):
    queue = deque()
    queue.append((start[0], start[1], None, [start]))
    seen = {}
    solutions = []
    while queue:
        r, c, last_dir, path = queue.popleft()
        if len(path) > 25:
            continue
        key = (r, c, last_dir, frozenset(path))
        if key in seen:
            continue
        seen[key] = True
        path_set = set(path)
        for dr, dc, dname in DIRS:
            if not allow_backward and last_dir and dname == OPPOSITE[last_dir]:
                continue
            mid_r, mid_c = r + dr, c + dc
            new_r, new_c = r + 2*dr, c + 2*dc
            if not (0 <= mid_r < ROWS and 0 <= mid_c < COLS):
                continue
            if not (0 <= new_r < ROWS and 0 <= new_c < COLS):
                continue
            if check_intermediate and is_blue(mid_r, mid_c):
                continue
            if is_blue(new_r, new_c):
                continue
            if (new_r, new_c) in path_set:
                continue
            new_path = path + [(new_r, new_c)]
            if (new_r, new_c) == end:
                solutions.append(new_path)
                cells = [f"{chr(65+cc)}{rr+1}" for rr,cc in new_path]
                print(f"SOLUTION ({len(new_path)-1} turns): {cells}")
                if len(solutions) >= 3:
                    return solutions
            else:
                queue.append((new_r, new_c, dname, new_path))
    return solutions

print("\n=== no backward, check intermediate ===")
sols = bfs_solve(False, True)
if not sols:
    print("\n=== allow backward, check intermediate ===")
    sols = bfs_solve(True, True)
if not sols:
    print("\n=== no backward, skip intermediate ===")
    sols = bfs_solve(False, False)
if not sols:
    print("\n=== allow backward, skip intermediate ===")
    sols = bfs_solve(True, False)

if sols:
    path = sols[0]
    print("\nPath details:")
    for j, (r, c) in enumerate(path):
        clr = cell_color[(r,c)]
        print(f"  Turn {j}: {chr(65+c)}{r+1} color={clr}")
    if len(path) >= 12:
        r11, c11 = path[11]
        clr = cell_color[(r11,c11)]
        print(f"\nTURN 11 ANSWER: {chr(65+c11)}{r11+1}, color={clr[2:]}")
    else:
        print(f"Path only {len(path)-1} turns long!")
else:
    print("NO SOLUTION FOUND")
